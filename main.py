import discord
import os
from dotenv import load_dotenv
import random
from game import get_rows
import textdistance
import openpyxl

load_dotenv()

client = discord.Client()

@client.event
async def on_ready():
    print('We have logged in as {0.user}'.format(client))

@client.event
async def on_message(message):
    # if message.author == client.user:
    #     return

    # if message.content.startswith('$hello'):
    #     await message.channel.send('Hello!')
    if message.content.startswith('$help'):
        str = "Welcome to the Fate GO Game!\nFind out which servant you are\n\nRules of the game:\n1. Enter 6 numbers between 1-18\nFormat: '$play num1 num2 num3 num4 num5 num6'\n2. Take care the sum does not cross 80, else you will have to retry"
        await message.channel.send(str)

    if message.content.startswith('$play'):
        str = message.content.replace('$play', '')
        str = str.strip()
        str = str.split(' ')
        print(str)
        if len(str) != 6:
            await message.channel.send("Invalid input")
            return
        for i in str:
            if int(i) < 1 or int(i) > 18:
                await message.channel.send("Invalid input")
                return
        # await message.channel.send("Your number is: " + str[0] + " " + str[1] + " " + str[2] + " " + str[3] + " " + str[4] + " " + str[5])
        # await message.channel.send("The sum is: " + str(int(str[0]) + int(str[1]) + int(str[2]) + int(str[3]) + int(str[4]) + int(str[5])))
        if int(str[0]) + int(str[1]) + int(str[2]) + int(str[3]) + int(str[4]) + int(str[5]) > 80:
            await message.channel.send("You have to retry")
            return
        # await message.channel.send("You are " + str(int(str[0]) + int(str[1]) + int(str[2]) + int(str[3]) + int(str[4]) + int(str[5])) + "% likely to be " + str[0] + " " + str[1] + " " + str[2] + " " + str[3] + " " + str[4] + " " + str[5])

        #convert str[0] through str[5] to int and make a list out of them
        nums = []
        for i in str:
            nums.append(int(i))
        
        random.shuffle(nums)

        rows = get_rows()
        rowdict = {}
        i=0
        for row in rows:
            rowdict[i] = [row[0], [row[1], row[2], row[3], row[4], row[5], row[6]], 100]
            i+=1

        i=0
        min=100
        for row in rowdict:
            k = textdistance.levenshtein(nums, rowdict[i][1])
            rowdict[i][2] = k
            if(k<min):
                min = k
            i+=1

        i=0
        strn=""
        for row in rowdict:
            if(rowdict[i][2] == min):
                # print("\033[1;32mYou are " + rowdict[i][0] + "\033[0;0m")
                await message.channel.send("You are " + rowdict[i][0])
                strn = rowdict[i][0]
                min=101
            i+=1
        
        wb = openpyxl.load_workbook('fgo.xlsx')
        sheet = wb.active
        i=0
        for row in sheet.iter_rows():
            if(row[0].value == strn):
                text = "Learn more about {ser} here"
                str1 = text.format(ser=strn)
                str2 = "<"+sheet.cell(row=(i+1), column=1).hyperlink.target+">"
                await message.channel.send(str1)
                await message.channel.send(str2)
            i+=1



token = os.getenv('TOKEN')

client.run(token)
